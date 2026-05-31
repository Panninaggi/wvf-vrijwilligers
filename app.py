from flask import (Flask, render_template, request, redirect, url_for,
                   session, flash, jsonify)
from werkzeug.security import check_password_hash

def generate_password_hash(pw):
    from werkzeug.security import generate_password_hash as _gph
    return _gph(pw, method='pbkdf2:sha256')
from functools import wraps
import sqlite3, os, json, smtplib, secrets
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'wvf-dev-secret-vervang-in-productie')
DB_PATH = os.environ.get('DB_PATH', os.path.join(os.path.dirname(__file__), 'vrijwilligers.db'))

# ── Email configuratie ─────────────────────────────────────────────────────────
# Stel in via omgevingsvariabelen of pas hieronder direct aan
EMAIL_HOST = os.environ.get('SMTP_HOST', '')
EMAIL_PORT = int(os.environ.get('SMTP_PORT', '587'))
EMAIL_USER = os.environ.get('SMTP_USER', '')
EMAIL_PASS = os.environ.get('SMTP_PASS', '')
EMAIL_FROM = os.environ.get('SMTP_FROM', 'WVF Vrijwilligers <noreply@wvf.nl>')

# ── Profiel clusters ───────────────────────────────────────────────────────────
PROFIELEN_SEED = [
    'Accommodatie', 'Activiteiten & Clubbinding', 'Administratie & Ondersteuning',
    'Arbitrage', 'Bestuur & Commissies', 'Communicatie & Media', 'Evenementen',
    'Financiën', 'Gastvrijheid & Ontvangst', 'Horeca', 'Jeugd', 'Senioren',
    'Sponsoring & Netwerk', 'Voetbalontwikkeling', 'Zorg & Veiligheid',
]

PROFIEL_CLUSTERS = [
    ('Sportief & Technisch',      ['Arbitrage', 'Jeugd', 'Senioren', 'Voetbalontwikkeling']),
    ('Evenementen & Hospitality', ['Accommodatie', 'Evenementen', 'Gastvrijheid & Ontvangst', 'Horeca']),
    ('Communicatie & PR',         ['Activiteiten & Clubbinding', 'Communicatie & Media', 'Sponsoring & Netwerk']),
    ('Bestuur & Organisatie',     ['Administratie & Ondersteuning', 'Bestuur & Commissies', 'Financiën', 'Zorg & Veiligheid']),
]

ALLE_ROLLEN = ['beheerder', 'roleigenaar', 'vrijwilliger']


# ── Database laag ─────────────────────────────────────────────────────────────
# Ondersteunt SQLite (lokaal) én PostgreSQL (Vercel/Neon) via dezelfde API.

DATABASE_URL = os.environ.get('DATABASE_URL', '')

_SCHEMA = '''
    CREATE TABLE IF NOT EXISTS vrijwilligers (
        id {auto}, naam TEXT, adres TEXT, email TEXT, telefoonnummer TEXT,
        profielen TEXT, aangemaakt TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    );
    CREATE TABLE IF NOT EXISTS eigenaren (
        id {auto}, voornaam TEXT NOT NULL, achternaam TEXT NOT NULL,
        email TEXT, telefoonnummer TEXT, aangemaakt TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    );
    CREATE TABLE IF NOT EXISTS profielen (
        id {auto}, naam TEXT NOT NULL UNIQUE,
        eigenaar_id INTEGER REFERENCES eigenaren(id) ON DELETE SET NULL,
        aangemaakt TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    );
    CREATE TABLE IF NOT EXISTS gebruikers (
        id {auto}, voornaam TEXT NOT NULL, achternaam TEXT NOT NULL,
        email TEXT NOT NULL UNIQUE, wachtwoord TEXT NOT NULL,
        actief INTEGER NOT NULL DEFAULT 1, aangemaakt TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    );
    CREATE TABLE IF NOT EXISTS gebruiker_rollen (
        gebruiker_id INTEGER NOT NULL REFERENCES gebruikers(id) ON DELETE CASCADE,
        rol TEXT NOT NULL, PRIMARY KEY (gebruiker_id, rol)
    );
    CREATE TABLE IF NOT EXISTS taken (
        id {auto},
        vrijwilliger_id INTEGER REFERENCES vrijwilligers(id) ON DELETE CASCADE,
        eigenaar_id INTEGER REFERENCES eigenaren(id) ON DELETE SET NULL,
        profiel TEXT, type TEXT NOT NULL DEFAULT 'intake',
        status TEXT NOT NULL DEFAULT 'Nieuw', opmerkingen TEXT,
        aangemaakt TIMESTAMP DEFAULT CURRENT_TIMESTAMP, bijgewerkt TIMESTAMP
    );
    CREATE TABLE IF NOT EXISTS intakes (
        id {auto}, taak_id INTEGER UNIQUE REFERENCES taken(id) ON DELETE CASCADE,
        vrijwilliger_id INTEGER REFERENCES vrijwilligers(id) ON DELETE CASCADE,
        formulier_data TEXT, status TEXT NOT NULL DEFAULT 'Concept',
        ingevuld TIMESTAMP DEFAULT CURRENT_TIMESTAMP, bijgewerkt TIMESTAMP
    );
    CREATE TABLE IF NOT EXISTS wachtwoord_tokens (
        token TEXT PRIMARY KEY,
        gebruiker_id INTEGER NOT NULL REFERENCES gebruikers(id) ON DELETE CASCADE,
        aangemaakt TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    );
'''


class _SQLiteConn:
    """SQLite-verbinding met dezelfde API als _PGConn."""

    def __init__(self, c):
        self._c = c

    def execute(self, sql, params=()):
        return self._c.execute(sql, params)

    def executemany(self, sql, params):
        return self._c.executemany(sql, params)

    def insert(self, sql, params=()):
        """INSERT en geeft het nieuwe id terug."""
        return self._c.execute(sql, params).lastrowid

    def col_exists(self, table, col):
        return col in [r[1] for r in self._c.execute(f'PRAGMA table_info({table})').fetchall()]

    def add_col(self, table, col, typ):
        if not self.col_exists(table, col):
            self._c.execute(f'ALTER TABLE {table} ADD COLUMN {col} {typ}')

    def scalar(self, sql, params=()):
        """Eerste kolom van eerste rij (voor COUNT queries)."""
        row = self._c.execute(sql, params).fetchone()
        return row[0] if row else 0

    def create_schema(self):
        self._c.executescript(_SCHEMA.format(auto='INTEGER PRIMARY KEY AUTOINCREMENT'))

    def commit(self):   self._c.commit()
    def close(self):    self._c.close()
    def rollback(self): pass


class _PGConn:
    """PostgreSQL-verbinding (psycopg2) met dezelfde API als _SQLiteConn."""

    def __init__(self, c):
        self._c = c

    def _fix(self, sql):
        import re
        sql = sql.replace('?', '%s')
        # SQLite :naam → PostgreSQL %(naam)s
        sql = re.sub(r':([a-zA-Z_][a-zA-Z0-9_]*)', r'%(\1)s', sql)
        return sql

    def execute(self, sql, params=()):
        cur = self._c.cursor()
        cur.execute(self._fix(sql), params or None)
        return cur

    def executemany(self, sql, params):
        import psycopg2.extras
        cur = self._c.cursor()
        psycopg2.extras.execute_batch(cur, self._fix(sql), params)
        return cur

    def insert(self, sql, params=()):
        cur = self._c.cursor()
        cur.execute(self._fix(sql) + ' RETURNING id', params or None)
        return cur.fetchone()['id']

    def col_exists(self, table, col):
        cur = self._c.cursor()
        cur.execute(
            "SELECT 1 FROM information_schema.columns WHERE table_name=%s AND column_name=%s",
            (table, col)
        )
        return bool(cur.fetchone())

    def add_col(self, table, col, typ):
        if not self.col_exists(table, col):
            self.execute(f'ALTER TABLE {table} ADD COLUMN {col} {typ}')

    def scalar(self, sql, params=()):
        cur = self._c.cursor()
        cur.execute(self._fix(sql), params or None)
        row = cur.fetchone()
        return list(row.values())[0] if row else 0

    def create_schema(self):
        cur = self._c.cursor()
        for stmt in _SCHEMA.format(auto='SERIAL PRIMARY KEY').strip().split(';'):
            stmt = stmt.strip()
            if stmt:
                cur.execute(stmt)

    def commit(self):   self._c.commit()
    def close(self):    self._c.close()
    def rollback(self): self._c.rollback()


def get_db():
    if DATABASE_URL:
        import re, urllib.parse
        url = DATABASE_URL.replace('postgres://', 'postgresql://', 1)
        url = re.sub(r'[&?]channel_binding=[^&]+', '', url)
        try:
            import psycopg2, psycopg2.extras
            raw = psycopg2.connect(url, cursor_factory=psycopg2.extras.RealDictCursor)
            return _PGConn(raw)
        except Exception:
            # Fallback: pg8000 (pure Python, werkt altijd op Vercel)
            import pg8000.dbapi
            p = urllib.parse.urlparse(url)
            raw = pg8000.dbapi.connect(
                host=p.hostname, port=p.port or 5432,
                database=p.path.lstrip('/'),
                user=p.username, password=p.password,
                ssl_context=True
            )
            raw.row_factory = pg8000.dbapi.DictRow
            return _PGConn(raw)
    raw = sqlite3.connect(DB_PATH)
    raw.row_factory = sqlite3.Row
    raw.execute('PRAGMA foreign_keys = ON')
    return _SQLiteConn(raw)


def init_db():
    conn = get_db()
    conn.create_schema()

    for col, typ in [
        ('gearchiveerd','INTEGER'),('archief_reden','TEXT'),('archief_datum','TIMESTAMP'),
        ('voornaam','TEXT'),('tussenvoegsel','TEXT'),('achternaam','TEXT'),
        ('postcode','TEXT'),('woonplaats','TEXT'),('geboortedatum','TEXT'),
        ('knvb_lid','TEXT'),('relatienummer','TEXT'),('ouder_verzorger','TEXT'),
        ('naam_kind','TEXT'),('team_kind','TEXT'),('eigen_bedrijf','TEXT'),
        ('sponsor_interesse','TEXT'),('vriend_wvf','TEXT'),('avg_toestemming','TEXT'),
        ('opmerkingen','TEXT'),('status_vrijwilliger','TEXT'),
    ]:
        conn.add_col('vrijwilligers', col, typ)

    for col, typ in [('tussenvoegsel','TEXT'),('gebruiker_id','INTEGER')]:
        conn.add_col('eigenaren', col, typ)

    for col, typ in [
        ('vog_nodig','TEXT'),('gedragscode_vereist','TEXT'),('avg_akkoord_vereist','TEXT'),
        ('tweede_eigenaar_id','INTEGER'),('tweede_eigenaar_actief','INTEGER'),
    ]:
        conn.add_col('profielen', col, typ)

    if conn.scalar('SELECT COUNT(*) FROM profielen') == 0:
        conn.executemany(
            'INSERT INTO profielen (naam) VALUES (?) ON CONFLICT (naam) DO NOTHING',
            [(p,) for p in PROFIELEN_SEED]
        )

    if conn.scalar('SELECT COUNT(*) FROM gebruikers') == 0:
        ww = secrets.token_urlsafe(10)
        uid = conn.insert(
            'INSERT INTO gebruikers (voornaam, achternaam, email, wachtwoord) VALUES (?,?,?,?)',
            ('Admin', 'WVF', 'admin@wvf.nl', generate_password_hash(ww))
        )
        conn.execute('INSERT INTO gebruiker_rollen (gebruiker_id, rol) VALUES (?,?)', (uid, 'beheerder'))
        print(f'\n  ┌─────────────────────────────────────────────┐')
        print(f'  │  Eerste beheerder aangemaakt               │')
        print(f'  │  E-mail:     admin@wvf.nl                  │')
        print(f'  │  Wachtwoord: {ww:<32}│')
        print(f'  └─────────────────────────────────────────────┘\n')

    conn.commit()
    conn.close()


# ── Auth helpers ───────────────────────────────────────────────────────────────

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            flash('Log eerst in om verder te gaan.', 'warning')
            return redirect(url_for('login', next=request.path))
        return f(*args, **kwargs)
    return decorated


def rol_vereist(*rollen):
    def decorator(f):
        @wraps(f)
        def decorated(*args, **kwargs):
            if 'user_id' not in session:
                return redirect(url_for('login'))
            if not heeft_rol(session['user_id'], *rollen):
                flash('Je hebt geen toegang tot deze pagina.', 'error')
                return redirect(url_for('taken'))
            return f(*args, **kwargs)
        return decorated
    return decorator


def heeft_rol(user_id, *rollen):
    conn = get_db()
    user_rollen = {r['rol'] for r in conn.execute(
        'SELECT rol FROM gebruiker_rollen WHERE gebruiker_id = ?', (user_id,)
    ).fetchall()}
    conn.close()
    return bool(user_rollen & set(rollen))


@app.context_processor
def inject_user():
    if 'user_id' in session:
        conn = get_db()
        user = conn.execute(
            'SELECT * FROM gebruikers WHERE id = ?', (session['user_id'],)
        ).fetchone()
        rollen = {r['rol'] for r in conn.execute(
            'SELECT rol FROM gebruiker_rollen WHERE gebruiker_id = ?', (session['user_id'],)
        ).fetchall()}
        conn.close()
        if user:
            return {'huidig_user': user, 'huidig_rollen': rollen}
    return {'huidig_user': None, 'huidig_rollen': set()}


# ── Email ──────────────────────────────────────────────────────────────────────

def send_email(to_addr, subject, html_body):
    if not EMAIL_HOST or not EMAIL_USER:
        print(f'\n  [EMAIL — niet verzonden, SMTP niet geconfigureerd]')
        print(f'  Aan: {to_addr}')
        print(f'  Onderwerp: {subject}\n')
        return False
    try:
        msg = MIMEMultipart('alternative')
        msg['Subject'] = subject
        msg['From'] = EMAIL_FROM
        msg['To'] = to_addr
        msg.attach(MIMEText(html_body, 'html', 'utf-8'))
        with smtplib.SMTP(EMAIL_HOST, EMAIL_PORT) as smtp:
            smtp.ehlo()
            smtp.starttls()
            smtp.login(EMAIL_USER, EMAIL_PASS)
            smtp.send_message(msg)
        return True
    except Exception as e:
        print(f'  [EMAIL FOUT] {e}')
        return False


def email_eigenaar_notificatie(eigenaar_email, eigenaar_naam, vrijwilliger_naam, profiel, taak_id):
    base_url = os.environ.get('APP_URL', 'http://localhost:5000')
    intake_url = f'{base_url}/intake/{taak_id}'
    html = f'''
    <div style="font-family:sans-serif;max-width:600px;margin:0 auto">
      <div style="background:#0f3460;color:white;padding:1.5rem;border-radius:8px 8px 0 0">
        <h2 style="margin:0">Nieuwe vrijwilliger aangemeld</h2>
      </div>
      <div style="background:#f8fafc;padding:1.5rem;border-radius:0 0 8px 8px;border:1px solid #e2e8f0">
        <p>Beste {eigenaar_naam},</p>
        <p><strong>{vrijwilliger_naam}</strong> heeft zich aangemeld voor het profiel
           <strong>{profiel}</strong>.</p>
        <p>Er staat een intaketaak voor je klaar. Vul het intakeformulier in om de aanmelding
           te verwerken.</p>
        <p style="margin-top:1.5rem">
          <a href="{intake_url}"
             style="background:#0f3460;color:white;padding:.75rem 1.5rem;border-radius:6px;
                    text-decoration:none;font-weight:600">
            Intakeformulier invullen →
          </a>
        </p>
        <p style="color:#94a3b8;font-size:.85rem;margin-top:2rem">
          WVF Vrijwilligersbeheer · <a href="{base_url}">{base_url}</a>
        </p>
      </div>
    </div>'''
    send_email(eigenaar_email, f'Nieuwe aanmelding: {vrijwilliger_naam} — {profiel}', html)


def email_welkom(to_addr, naam, wachtwoord):
    base_url = os.environ.get('APP_URL', 'http://localhost:5000')
    html = f'''
    <div style="font-family:sans-serif;max-width:600px;margin:0 auto">
      <div style="background:#1A6CC4;color:white;padding:1.5rem;border-radius:8px 8px 0 0">
        <h2 style="margin:0">Welkom bij WVF Vrijwilligersbeheer</h2>
      </div>
      <div style="background:#f8fafc;padding:1.5rem;border-radius:0 0 8px 8px;border:1px solid #e2e8f0">
        <p>Beste {naam},</p>
        <p>Er is een account voor je aangemaakt. Je kunt inloggen met:</p>
        <table style="margin:1rem 0;border-collapse:collapse">
          <tr><td style="padding:.3rem 1rem .3rem 0;color:#64748b">E-mail</td><td><strong>{to_addr}</strong></td></tr>
          <tr><td style="padding:.3rem 1rem .3rem 0;color:#64748b">Wachtwoord</td><td><strong>{wachtwoord}</strong></td></tr>
        </table>
        <p style="margin-top:1.5rem">
          <a href="{base_url}/login"
             style="background:#1A6CC4;color:white;padding:.75rem 1.5rem;border-radius:6px;text-decoration:none;font-weight:600">
            Inloggen →
          </a>
        </p>
        <p style="color:#94a3b8;font-size:.85rem;margin-top:2rem">
          Wijzig je wachtwoord na je eerste inlog via het menu rechtsboven.
        </p>
      </div>
    </div>'''
    send_email(to_addr, 'Je account voor WVF Vrijwilligersbeheer', html)


def email_wachtwoord_reset(to_addr, naam, token):
    base_url = os.environ.get('APP_URL', 'http://localhost:5000')
    html = f'''
    <div style="font-family:sans-serif;max-width:600px;margin:0 auto">
      <div style="background:#1A6CC4;color:white;padding:1.5rem;border-radius:8px 8px 0 0">
        <h2 style="margin:0">Wachtwoord opnieuw instellen</h2>
      </div>
      <div style="background:#f8fafc;padding:1.5rem;border-radius:0 0 8px 8px;border:1px solid #e2e8f0">
        <p>Beste {naam},</p>
        <p>Klik op de knop hieronder om een nieuw wachtwoord in te stellen. De link is <strong>1 uur geldig</strong>.</p>
        <p style="margin-top:1.5rem">
          <a href="{base_url}/wachtwoord-reset/{token}"
             style="background:#1A6CC4;color:white;padding:.75rem 1.5rem;border-radius:6px;text-decoration:none;font-weight:600">
            Wachtwoord opnieuw instellen →
          </a>
        </p>
        <p style="color:#94a3b8;font-size:.85rem;margin-top:2rem">
          Heb je dit niet aangevraagd? Dan hoef je niets te doen.
        </p>
      </div>
    </div>'''
    send_email(to_addr, 'Wachtwoord opnieuw instellen — WVF', html)


# ── QR-code ────────────────────────────────────────────────────────────────────

def genereer_qr_base64(url):
    try:
        import qrcode, io, base64
        qr = qrcode.QRCode(box_size=8, border=3,
                           error_correction=qrcode.constants.ERROR_CORRECT_M)
        qr.add_data(url)
        qr.make(fit=True)
        img = qr.make_image(fill_color='#0f3460', back_color='white')
        buf = io.BytesIO()
        img.save(buf, format='PNG')
        return base64.b64encode(buf.getvalue()).decode()
    except ImportError:
        return None


# ── Template filters ───────────────────────────────────────────────────────────

@app.template_filter('selectattr_any')
def selectattr_any_filter(d, keys):
    """Geeft True als tenminste één sleutel een waarde heeft in dict d."""
    return any(d.get(k) for k in keys)


@app.template_filter('datum')
def datum_filter(waarde):
    """Werkt met zowel SQLite strings als PostgreSQL datetime objecten."""
    if waarde is None:
        return ''
    if hasattr(waarde, 'strftime'):
        return waarde.strftime('%d-%m-%Y')
    return str(waarde)[:10]


@app.template_filter('display_naam')
def display_naam_filter(v):
    if v['voornaam']:
        return ' '.join(p for p in [v['voornaam'], v['tussenvoegsel'], v['achternaam']] if p)
    return v['naam'] or ''


def cluster_profielen(profielen_rows):
    alle = [r['naam'] for r in profielen_rows]
    naam_set = set(alle)
    result, gebruikt = [], set()
    for cluster, namen in PROFIEL_CLUSTERS:
        items = [n for n in namen if n in naam_set]
        if items:
            result.append((cluster, items))
            gebruikt.update(items)
    overig = [n for n in alle if n not in gebruikt]
    if overig:
        result.append(('Overig', overig))
    return result


# ── Auth routes ────────────────────────────────────────────────────────────────

@app.errorhandler(500)
def fout500(e):
    import traceback
    return f'<pre style="font-size:13px">{traceback.format_exc()}</pre>', 500


@app.route('/healthz')
def healthz():
    """Diagnosepagina — verwijder na productie."""
    try:
        conn = get_db()
        n = conn.scalar('SELECT COUNT(*) FROM gebruikers')
        conn.close()
        return f'OK — database bereikbaar, {n} gebruiker(s)'
    except Exception as e:
        return f'FOUT: {e}', 500


@app.route('/login', methods=['GET', 'POST'])
def login():
    if 'user_id' in session:
        return redirect(url_for('taken'))
    if request.method == 'POST':
        email = request.form.get('email', '').strip().lower()
        wachtwoord = request.form.get('wachtwoord', '')
        conn = get_db()
        user = conn.execute(
            'SELECT * FROM gebruikers WHERE LOWER(email) = ? AND actief = 1', (email,)
        ).fetchone()
        conn.close()
        if user and check_password_hash(user['wachtwoord'], wachtwoord):
            session.permanent = True
            session['user_id'] = user['id']
            next_url = request.args.get('next') or url_for('taken')
            return redirect(next_url)
        flash('Onbekend e-mailadres of onjuist wachtwoord.', 'error')
    return render_template('login.html')


@app.route('/logout')
def logout():
    session.clear()
    flash('Je bent uitgelogd.', 'info')
    return redirect(url_for('login'))


@app.route('/wachtwoord-vergeten', methods=['GET', 'POST'])
def wachtwoord_vergeten():
    if request.method == 'POST':
        email = request.form.get('email', '').strip().lower()
        conn = get_db()
        user = conn.execute(
            'SELECT * FROM gebruikers WHERE LOWER(email) = ? AND actief = 1', (email,)
        ).fetchone()
        if user:
            token = secrets.token_urlsafe(32)
            conn.execute(
                'INSERT INTO wachtwoord_tokens (token, gebruiker_id) VALUES (?,?)',
                (token, user['id'])
            )
            conn.commit()
            naam = f"{user['voornaam']} {user['achternaam']}"
            email_wachtwoord_reset(email, naam, token)
        conn.close()
        # Altijd dezelfde melding (voorkomt dat je kunt achterhalen of een e-mail bestaat)
        flash('Als dit e-mailadres bekend is, ontvang je een e-mail met een resetlink.', 'info')
        return redirect(url_for('wachtwoord_vergeten'))
    return render_template('wachtwoord_vergeten.html')


@app.route('/wachtwoord-reset/<token>', methods=['GET', 'POST'])
def wachtwoord_reset(token):
    conn = get_db()
    record = conn.execute('''
        SELECT wt.*, g.voornaam, g.achternaam, g.email
        FROM wachtwoord_tokens wt
        JOIN gebruikers g ON wt.gebruiker_id = g.id
        WHERE wt.token = ?
    ''', (token,)).fetchone()

    if not record:
        conn.close()
        flash('Ongeldige of verlopen resetlink.', 'error')
        return redirect(url_for('login'))

    # Token verlopen na 1 uur
    from datetime import datetime, timezone, timedelta
    aangemaakt = record['aangemaakt']
    if hasattr(aangemaakt, 'replace'):
        aangemaakt = aangemaakt.replace(tzinfo=None)
    else:
        try:
            aangemaakt = datetime.fromisoformat(str(aangemaakt))
        except Exception:
            aangemaakt = datetime.now()
    if datetime.now() - aangemaakt > timedelta(hours=1):
        conn.execute('DELETE FROM wachtwoord_tokens WHERE token = ?', (token,))
        conn.commit()
        conn.close()
        flash('Deze resetlink is verlopen. Vraag een nieuwe aan.', 'error')
        return redirect(url_for('wachtwoord_vergeten'))

    if request.method == 'POST':
        nieuw = request.form.get('wachtwoord', '')
        bevestiging = request.form.get('bevestiging', '')
        if len(nieuw) < 6:
            flash('Wachtwoord moet minimaal 6 tekens zijn.', 'error')
            conn.close()
            return render_template('wachtwoord_reset.html', token=token, naam=record['voornaam'])
        if nieuw != bevestiging:
            flash('De wachtwoorden komen niet overeen.', 'error')
            conn.close()
            return render_template('wachtwoord_reset.html', token=token, naam=record['voornaam'])
        conn.execute('UPDATE gebruikers SET wachtwoord = ? WHERE id = ?',
                     (generate_password_hash(nieuw), record['gebruiker_id']))
        conn.execute('DELETE FROM wachtwoord_tokens WHERE token = ?', (token,))
        conn.commit()
        conn.close()
        flash('Wachtwoord gewijzigd. Je kunt nu inloggen.', 'success')
        return redirect(url_for('login'))

    conn.close()
    return render_template('wachtwoord_reset.html', token=token, naam=record['voornaam'])


@app.route('/mijn-wachtwoord', methods=['GET', 'POST'])
@login_required
def mijn_wachtwoord():
    if request.method == 'POST':
        huidig = request.form.get('huidig', '')
        nieuw = request.form.get('nieuw', '')
        bevestiging = request.form.get('bevestiging', '')
        conn = get_db()
        user = conn.execute('SELECT * FROM gebruikers WHERE id = ?',
                            (session['user_id'],)).fetchone()
        if not check_password_hash(user['wachtwoord'], huidig):
            conn.close()
            flash('Huidig wachtwoord klopt niet.', 'error')
            return render_template('mijn_wachtwoord.html')
        if len(nieuw) < 6:
            conn.close()
            flash('Nieuw wachtwoord moet minimaal 6 tekens zijn.', 'error')
            return render_template('mijn_wachtwoord.html')
        if nieuw != bevestiging:
            conn.close()
            flash('De wachtwoorden komen niet overeen.', 'error')
            return render_template('mijn_wachtwoord.html')
        conn.execute('UPDATE gebruikers SET wachtwoord = ? WHERE id = ?',
                     (generate_password_hash(nieuw), session['user_id']))
        conn.commit()
        conn.close()
        flash('Wachtwoord succesvol gewijzigd.', 'success')
        return redirect(url_for('taken'))
    return render_template('mijn_wachtwoord.html')


# ── Publieke registratie ───────────────────────────────────────────────────────

@app.route('/registreren', methods=['GET'])
def registreren():
    conn = get_db()
    profielen_rows = conn.execute('SELECT naam FROM profielen ORDER BY naam').fetchall()
    conn.close()
    return render_template('registreren.html',
                           profielen_clusters=cluster_profielen(profielen_rows))


@app.route('/registreren', methods=['POST'])
def registreren_post():
    voornaam = request.form.get('voornaam', '').strip()
    tussenvoegsel = request.form.get('tussenvoegsel', '').strip()
    achternaam = request.form.get('achternaam', '').strip()
    naam = ' '.join(p for p in [voornaam, tussenvoegsel, achternaam] if p)
    if not naam:
        flash('Vul minimaal een naam in.', 'error')
        return redirect(url_for('registreren'))

    geselecteerde_profielen = request.form.getlist('profielen')

    conn = get_db()
    vw_id = conn.insert('''
        INSERT INTO vrijwilligers
        (naam, voornaam, tussenvoegsel, achternaam, adres, postcode, woonplaats,
         geboortedatum, email, telefoonnummer, profielen, status_vrijwilliger)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?)
    ''', (naam, voornaam, tussenvoegsel, achternaam,
          request.form.get('adres','').strip(),
          request.form.get('postcode','').strip(),
          request.form.get('woonplaats','').strip(),
          request.form.get('geboortedatum','').strip(),
          request.form.get('email','').strip(),
          request.form.get('telefoonnummer','').strip(),
          '||'.join(geselecteerde_profielen),
          'Nieuw'))

    # Taak + email per profiel
    for profiel_naam in geselecteerde_profielen:
        profiel_row = conn.execute('''
            SELECT p.id, p.eigenaar_id,
                   p.tweede_eigenaar_id, p.tweede_eigenaar_actief,
                   e.email  AS eigenaar_email,
                   e.voornaam||COALESCE(' '||NULLIF(e.tussenvoegsel,''),'')||' '||e.achternaam AS eigenaar_naam,
                   e2.email AS tweede_email,
                   e2.voornaam||COALESCE(' '||NULLIF(e2.tussenvoegsel,''),'')||' '||e2.achternaam AS tweede_naam
            FROM profielen p
            LEFT JOIN eigenaren e  ON p.eigenaar_id        = e.id
            LEFT JOIN eigenaren e2 ON p.tweede_eigenaar_id = e2.id
            WHERE p.naam = ?
        ''', (profiel_naam,)).fetchone()

        eigenaar_id = profiel_row['eigenaar_id'] if profiel_row else None
        taak_id = conn.insert(
            'INSERT INTO taken (vrijwilliger_id, eigenaar_id, profiel, type, status) VALUES (?,?,?,?,?)',
            (vw_id, eigenaar_id, profiel_naam, 'intake', 'Nieuw')
        )

        if profiel_row and profiel_row['eigenaar_email']:
            email_eigenaar_notificatie(
                profiel_row['eigenaar_email'],
                profiel_row['eigenaar_naam'],
                naam, profiel_naam, taak_id
            )

        # Tweede bevoegde — alleen als actief
        if profiel_row and profiel_row['tweede_eigenaar_actief'] and profiel_row['tweede_email']:
            email_eigenaar_notificatie(
                profiel_row['tweede_email'],
                profiel_row['tweede_naam'],
                naam, profiel_naam, taak_id
            )

    conn.commit()
    conn.close()
    return redirect(url_for('registreren_succes'))


@app.route('/registreren/succes')
def registreren_succes():
    return render_template('registreren_succes.html')


@app.route('/qr')
@login_required
def qr_code():
    base_url = request.host_url.rstrip('/')
    reg_url = f'{base_url}/registreren'
    qr_data = genereer_qr_base64(reg_url)
    return render_template('qr.html', reg_url=reg_url, qr_data=qr_data)


# ── Vrijwilligers (admin) ──────────────────────────────────────────────────────

@app.route('/')
@login_required
def index():
    if not heeft_rol(session['user_id'], 'beheerder'):
        return redirect(url_for('taken'))
    zoek = request.args.get('q', '').strip()
    conn = get_db()

    taak_subquery = '''
        LEFT JOIN (
            SELECT vrijwilliger_id,
                   COUNT(*) AS taken_totaal,
                   SUM(CASE WHEN status = 'Voltooid'        THEN 1 ELSE 0 END) AS taken_voltooid,
                   SUM(CASE WHEN status = 'In behandeling'  THEN 1 ELSE 0 END) AS taken_lopend,
                   SUM(CASE WHEN status = 'Nieuw'           THEN 1 ELSE 0 END) AS taken_nieuw
            FROM taken GROUP BY vrijwilliger_id
        ) t ON t.vrijwilliger_id = v.id
    '''

    actief_filter = "AND (v.gearchiveerd IS NULL OR v.gearchiveerd = 0)"
    if zoek:
        q = f'%{zoek}%'
        rows = conn.execute(
            f'''SELECT v.*, t.taken_totaal, t.taken_voltooid, t.taken_lopend, t.taken_nieuw
                FROM vrijwilligers v {taak_subquery}
                WHERE (v.naam LIKE ? OR v.voornaam LIKE ? OR v.achternaam LIKE ?
                   OR v.email LIKE ? OR v.profielen LIKE ? OR v.woonplaats LIKE ?)
                {actief_filter}
                ORDER BY v.achternaam, v.voornaam, v.naam''', (q,q,q,q,q,q)
        ).fetchall()
    else:
        rows = conn.execute(
            f'''SELECT v.*, t.taken_totaal, t.taken_voltooid, t.taken_lopend, t.taken_nieuw
                FROM vrijwilligers v {taak_subquery}
                WHERE (v.gearchiveerd IS NULL OR v.gearchiveerd = 0)
                ORDER BY v.achternaam, v.voornaam, v.naam'''
        ).fetchall()
    conn.close()
    return render_template('index.html', vrijwilligers=rows, zoek=zoek)


@app.route('/toevoegen', methods=['GET'])
@login_required
@rol_vereist('beheerder')
def toevoegen_form():
    conn = get_db()
    profielen_rows = conn.execute('SELECT naam FROM profielen ORDER BY naam').fetchall()
    conn.close()
    return render_template('toevoegen.html', profielen_clusters=cluster_profielen(profielen_rows))


@app.route('/toevoegen', methods=['POST'])
@login_required
@rol_vereist('beheerder')
def toevoegen():
    voornaam = request.form.get('voornaam','').strip()
    tussenvoegsel = request.form.get('tussenvoegsel','').strip()
    achternaam = request.form.get('achternaam','').strip()
    naam = ' '.join(p for p in [voornaam, tussenvoegsel, achternaam] if p)
    if not naam:
        return redirect(url_for('toevoegen_form'))
    conn = get_db()
    conn.execute('''
        INSERT INTO vrijwilligers
        (naam,voornaam,tussenvoegsel,achternaam,adres,postcode,woonplaats,
         geboortedatum,email,telefoonnummer,knvb_lid,relatienummer,
         ouder_verzorger,naam_kind,team_kind,eigen_bedrijf,
         sponsor_interesse,vriend_wvf,avg_toestemming,opmerkingen,
         status_vrijwilliger,profielen)
        VALUES
        (:naam,:voornaam,:tussenvoegsel,:achternaam,:adres,:postcode,:woonplaats,
         :geboortedatum,:email,:telefoonnummer,:knvb_lid,:relatienummer,
         :ouder_verzorger,:naam_kind,:team_kind,:eigen_bedrijf,
         :sponsor_interesse,:vriend_wvf,:avg_toestemming,:opmerkingen,
         :status_vrijwilliger,:profielen)
    ''', {
        'naam': naam, 'voornaam': voornaam, 'tussenvoegsel': tussenvoegsel,
        'achternaam': achternaam,
        'adres': request.form.get('adres','').strip(),
        'postcode': request.form.get('postcode','').strip(),
        'woonplaats': request.form.get('woonplaats','').strip(),
        'geboortedatum': request.form.get('geboortedatum','').strip(),
        'email': request.form.get('email','').strip(),
        'telefoonnummer': request.form.get('telefoonnummer','').strip(),
        'knvb_lid': request.form.get('knvb_lid',''),
        'relatienummer': request.form.get('relatienummer','').strip(),
        'ouder_verzorger': request.form.get('ouder_verzorger',''),
        'naam_kind': request.form.get('naam_kind','').strip(),
        'team_kind': request.form.get('team_kind','').strip(),
        'eigen_bedrijf': request.form.get('eigen_bedrijf',''),
        'sponsor_interesse': request.form.get('sponsor_interesse',''),
        'vriend_wvf': request.form.get('vriend_wvf',''),
        'avg_toestemming': request.form.get('avg_toestemming',''),
        'opmerkingen': request.form.get('opmerkingen','').strip(),
        'status_vrijwilliger': request.form.get('status_vrijwilliger',''),
        'profielen': '||'.join(request.form.getlist('profielen')),
    })
    conn.commit()
    conn.close()
    return redirect(url_for('index'))


@app.route('/vrijwilligers/<int:vid>', methods=['GET'])
@login_required
@rol_vereist('beheerder')
def vrijwilliger_detail(vid):
    conn = get_db()
    v = conn.execute('SELECT * FROM vrijwilligers WHERE id = ?', (vid,)).fetchone()
    if not v:
        conn.close()
        flash('Vrijwilliger niet gevonden.', 'error')
        return redirect(url_for('index'))
    profielen_rows = conn.execute('SELECT naam FROM profielen ORDER BY naam').fetchall()
    intakes_rows = conn.execute('''
        SELECT t.profiel, i.formulier_data, i.status,
               i.ingevuld, i.bijgewerkt
        FROM intakes i
        JOIN taken t ON i.taak_id = t.id
        WHERE t.vrijwilliger_id = ?
        ORDER BY t.profiel
    ''', (vid,)).fetchall()
    conn.close()

    # Parseer JSON per intake
    intakes = []
    for row in intakes_rows:
        data = json.loads(row['formulier_data']) if row['formulier_data'] else {}
        intakes.append({'profiel': row['profiel'], 'status': row['status'],
                        'ingevuld': row['ingevuld'], 'data': data})

    return render_template('vrijwilliger_detail.html',
                           v=v,
                           profielen_clusters=cluster_profielen(profielen_rows),
                           intakes=intakes)


@app.route('/vrijwilligers/<int:vid>', methods=['POST'])
@login_required
@rol_vereist('beheerder')
def vrijwilliger_opslaan(vid):
    voornaam = request.form.get('voornaam', '').strip()
    tussenvoegsel = request.form.get('tussenvoegsel', '').strip()
    achternaam = request.form.get('achternaam', '').strip()
    naam = ' '.join(p for p in [voornaam, tussenvoegsel, achternaam] if p)
    if not naam:
        flash('Voornaam en achternaam zijn verplicht.', 'error')
        return redirect(url_for('vrijwilliger_detail', vid=vid))

    list_fields = {'profielen'}
    data = {}
    for key in request.form:
        if key in list_fields:
            data[key] = '||'.join(request.form.getlist(key))
        else:
            data[key] = request.form.get(key, '').strip()

    conn = get_db()
    conn.execute('''UPDATE vrijwilligers SET
        naam=:naam, voornaam=:voornaam, tussenvoegsel=:tussenvoegsel, achternaam=:achternaam,
        adres=:adres, postcode=:postcode, woonplaats=:woonplaats, geboortedatum=:geboortedatum,
        email=:email, telefoonnummer=:telefoonnummer, knvb_lid=:knvb_lid,
        relatienummer=:relatienummer, ouder_verzorger=:ouder_verzorger,
        naam_kind=:naam_kind, team_kind=:team_kind, eigen_bedrijf=:eigen_bedrijf,
        sponsor_interesse=:sponsor_interesse, vriend_wvf=:vriend_wvf,
        avg_toestemming=:avg_toestemming, opmerkingen=:opmerkingen,
        status_vrijwilliger=:status_vrijwilliger, profielen=:profielen
        WHERE id=:id''', {**data, 'naam': naam, 'id': vid})
    conn.commit()
    conn.close()
    flash('Gegevens opgeslagen.', 'success')
    return redirect(url_for('vrijwilliger_detail', vid=vid))


@app.route('/archiveren/<int:vid>', methods=['POST'])
@login_required
@rol_vereist('beheerder')
def archiveren(vid):
    reden = request.form.get('reden', '').strip()
    conn = get_db()
    conn.execute('''UPDATE vrijwilligers
                    SET gearchiveerd = 1, archief_reden = ?, archief_datum = CURRENT_TIMESTAMP
                    WHERE id = ?''', (reden, vid))
    conn.commit()
    conn.close()
    flash('Vrijwilliger gearchiveerd.', 'info')
    return redirect(url_for('index'))


@app.route('/herstellen/<int:vid>', methods=['POST'])
@login_required
@rol_vereist('beheerder')
def herstellen(vid):
    conn = get_db()
    conn.execute('''UPDATE vrijwilligers
                    SET gearchiveerd = 0, archief_reden = NULL, archief_datum = NULL
                    WHERE id = ?''', (vid,))
    conn.commit()
    conn.close()
    flash('Vrijwilliger hersteld naar actief.', 'success')
    return redirect(url_for('archief'))


@app.route('/archief')
@login_required
@rol_vereist('beheerder')
def archief():
    conn = get_db()
    rows = conn.execute('''
        SELECT * FROM vrijwilligers
        WHERE gearchiveerd = 1
        ORDER BY archief_datum DESC
    ''').fetchall()
    conn.close()
    return render_template('archief.html', vrijwilligers=rows)


# ── Excel import ──────────────────────────────────────────────────────────────

IMPORT_KOLOMMEN = [
    'Voornaam', 'Tussenvoegsel', 'Achternaam', 'Adres', 'Postcode',
    'Woonplaats', 'Geboortedatum', 'E-mailadres', 'Telefoonnummer', 'Profielen',
]


@app.route('/import', methods=['GET'])
@login_required
@rol_vereist('beheerder')
def import_form():
    return render_template('import.html')


@app.route('/import/template')
@login_required
@rol_vereist('beheerder')
def import_template():
    import openpyxl, io
    from flask import send_file
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Vrijwilligers'

    # Headers met opmaak
    from openpyxl.styles import Font, PatternFill, Alignment
    header_fill = PatternFill('solid', fgColor='1A6CC4')
    for col, naam in enumerate(IMPORT_KOLOMMEN, 1):
        cel = ws.cell(row=1, column=col, value=naam)
        cel.font = Font(bold=True, color='FFFFFF')
        cel.fill = header_fill
        cel.alignment = Alignment(horizontal='center')
        ws.column_dimensions[cel.column_letter].width = max(14, len(naam) + 2)

    # Voorbeeldrij
    ws.append(['Jan', 'van', 'Janssen', 'Dorpstraat 1', '8000 AA',
               'Zwolle', '01-01-1985', 'jan@wvf.nl', '06-12345678',
               'Horeca, Jeugd'])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name='wvf-vrijwilligers-import.xlsx')


@app.route('/import', methods=['POST'])
@login_required
@rol_vereist('beheerder')
def import_verwerken():
    bestand = request.files.get('bestand')
    if not bestand or not bestand.filename.endswith(('.xlsx', '.xls')):
        flash('Selecteer een geldig Excel-bestand (.xlsx).', 'error')
        return redirect(url_for('import_form'))

    try:
        import openpyxl, io
        wb = openpyxl.load_workbook(io.BytesIO(bestand.read()), data_only=True)
        ws = wb.active
    except Exception as e:
        flash(f'Kon bestand niet lezen: {e}', 'error')
        return redirect(url_for('import_form'))

    # Kolomtoewijzing op basis van headernamen (hoofdlettersonafhankelijk)
    headers = {}
    for col in ws.iter_cols(min_row=1, max_row=1):
        for cel in col:
            if cel.value:
                headers[str(cel.value).strip().lower()] = cel.column - 1

    def kolom(rij_waarden, *namen):
        for naam in namen:
            idx = headers.get(naam.lower())
            if idx is not None and idx < len(rij_waarden):
                val = rij_waarden[idx]
                return str(val).strip() if val is not None else ''
        return ''

    conn = get_db()
    # Haal geldige profielnamen op
    geldige_profielen = {r['naam'].lower(): r['naam']
                         for r in conn.execute('SELECT naam FROM profielen').fetchall()}

    toegevoegd, overgeslagen, fouten = 0, 0, []

    for rijnr, rij in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if all(v is None or str(v).strip() == '' for v in rij):
            continue  # lege rij

        voornaam   = kolom(rij, 'voornaam')
        achternaam = kolom(rij, 'achternaam')

        if not voornaam or not achternaam:
            fouten.append(f'Rij {rijnr}: voornaam of achternaam ontbreekt — overgeslagen')
            overgeslagen += 1
            continue

        tussenvoegsel = kolom(rij, 'tussenvoegsel')
        naam = ' '.join(p for p in [voornaam, tussenvoegsel, achternaam] if p)

        # Profielen verwerken
        profiel_tekst = kolom(rij, 'profielen', 'profiel')
        profiel_namen = []
        for p in profiel_tekst.split(','):
            p = p.strip()
            if p.lower() in geldige_profielen:
                profiel_namen.append(geldige_profielen[p.lower()])
            elif p:
                fouten.append(f'Rij {rijnr}: onbekend profiel "{p}" — overgeslagen')

        try:
            conn.insert('''
                INSERT INTO vrijwilligers
                (naam, voornaam, tussenvoegsel, achternaam, adres, postcode,
                 woonplaats, geboortedatum, email, telefoonnummer,
                 profielen, status_vrijwilliger)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?)
            ''', (naam, voornaam, tussenvoegsel, achternaam,
                  kolom(rij, 'adres'),
                  kolom(rij, 'postcode'),
                  kolom(rij, 'woonplaats'),
                  kolom(rij, 'geboortedatum'),
                  kolom(rij, 'e-mailadres', 'email'),
                  kolom(rij, 'telefoonnummer', 'telefoon'),
                  '||'.join(profiel_namen),
                  'Nieuw'))
            toegevoegd += 1
        except Exception as e:
            fouten.append(f'Rij {rijnr} ({naam}): {e}')
            overgeslagen += 1

    conn.commit()
    conn.close()

    return render_template('import.html',
                           toegevoegd=toegevoegd,
                           overgeslagen=overgeslagen,
                           fouten=fouten,
                           klaar=True)


# ── Taken ──────────────────────────────────────────────────────────────────────

@app.route('/taken')
@login_required
def taken():
    user_id = session['user_id']
    conn = get_db()
    is_beheerder = heeft_rol(user_id, 'beheerder')

    eigenaar_sel = '''
        e.voornaam || COALESCE(' ' || NULLIF(e.tussenvoegsel,''), '')
        || ' ' || e.achternaam AS eigenaar_naam
    '''

    base_q = f'''
        SELECT t.*, v.naam AS vw_naam, v.voornaam AS vw_voornaam,
               v.tussenvoegsel AS vw_tv, v.achternaam AS vw_achternaam,
               v.email AS vw_email,
               {eigenaar_sel},
               i.id AS intake_id, i.status AS intake_status
        FROM taken t
        JOIN vrijwilligers v ON t.vrijwilliger_id = v.id
        LEFT JOIN eigenaren e ON t.eigenaar_id = e.id
        LEFT JOIN intakes i ON i.taak_id = t.id
    '''

    if is_beheerder:
        rows = conn.execute(base_q + ' ORDER BY t.aangemaakt DESC').fetchall()
    else:
        eigenaar = conn.execute(
            'SELECT id FROM eigenaren WHERE gebruiker_id = ?', (user_id,)
        ).fetchone()
        if eigenaar:
            rows = conn.execute(
                base_q + ' WHERE t.eigenaar_id = ? ORDER BY t.aangemaakt DESC',
                (eigenaar['id'],)
            ).fetchall()
        else:
            rows = []

    conn.close()
    open_taken    = [t for t in rows if t['status'] != 'Voltooid']
    voltooide     = [t for t in rows if t['status'] == 'Voltooid']
    return render_template('taken.html', open_taken=open_taken, voltooide=voltooide)


# ── Intake ─────────────────────────────────────────────────────────────────────

@app.route('/intake/<int:taak_id>', methods=['GET'])
@login_required
def intake_form(taak_id):
    conn = get_db()
    taak = conn.execute('''
        SELECT t.*, v.naam AS vw_naam, v.voornaam AS vw_voornaam,
               v.tussenvoegsel AS vw_tv, v.achternaam AS vw_achternaam,
               v.email AS vw_email, v.telefoonnummer AS vw_tel,
               v.geboortedatum AS vw_gbd, v.woonplaats AS vw_woonplaats
        FROM taken t JOIN vrijwilligers v ON t.vrijwilliger_id = v.id
        WHERE t.id = ?
    ''', (taak_id,)).fetchone()
    if not taak:
        flash('Taak niet gevonden.', 'error')
        return redirect(url_for('taken'))

    intake = conn.execute('SELECT * FROM intakes WHERE taak_id = ?', (taak_id,)).fetchone()
    formulier_data = json.loads(intake['formulier_data']) if intake and intake['formulier_data'] else {}

    # Haal medevrijwilligers op voor het partnerveld (Gastvrijheid & Ontvangst)
    partner_vrijwilligers = []
    if taak['profiel'] == 'Gastvrijheid & Ontvangst':
        partner_vrijwilligers = conn.execute(
            """SELECT id, naam, voornaam, tussenvoegsel, achternaam
               FROM vrijwilligers
               WHERE (profielen LIKE ? OR profielen LIKE ? OR profielen = ?)
               AND (gearchiveerd IS NULL OR gearchiveerd = 0)
               AND id != ?
               ORDER BY achternaam, voornaam, naam""",
            ('%||Gastvrijheid & Ontvangst%', 'Gastvrijheid & Ontvangst||%',
             'Gastvrijheid & Ontvangst', taak['vrijwilliger_id'])
        ).fetchall()

    conn.close()
    return render_template('intake.html', taak=taak, intake=intake,
                           data=formulier_data,
                           partner_vrijwilligers=partner_vrijwilligers)


@app.route('/intake/<int:taak_id>', methods=['POST'])
@login_required
def intake_opslaan(taak_id):
    conn = get_db()
    taak = conn.execute('SELECT * FROM taken WHERE id = ?', (taak_id,)).fetchone()
    if not taak:
        conn.close()
        flash('Taak niet gevonden.', 'error')
        return redirect(url_for('taken'))

    # Sla alle formuliervelden dynamisch op — werkt voor elk profiel
    list_fields = {'voorkeur_werkzaamheden', 'rollen_profiel', 'trainers_diploma', 'voorkeur_partners'}
    formulier_data = {}
    for key in request.form.keys():
        if key == 'actie':
            continue
        formulier_data[key] = (request.form.getlist(key)
                               if key in list_fields
                               else request.form.get(key, '').strip())
    for lf in list_fields:
        formulier_data.setdefault(lf, [])

    actie = request.form.get('actie', 'opslaan')
    intake_status = 'Ingediend' if actie == 'indienen' else 'Concept'

    existing = conn.execute('SELECT id FROM intakes WHERE taak_id = ?', (taak_id,)).fetchone()
    if existing:
        conn.execute('''UPDATE intakes SET formulier_data=?, status=?, bijgewerkt=CURRENT_TIMESTAMP
                        WHERE taak_id=?''',
                     (json.dumps(formulier_data, ensure_ascii=False), intake_status, taak_id))
    else:
        conn.execute('''INSERT INTO intakes (taak_id, vrijwilliger_id, formulier_data, status)
                        VALUES (?,?,?,?)''',
                     (taak_id, taak['vrijwilliger_id'],
                      json.dumps(formulier_data, ensure_ascii=False), intake_status))

    # Taak-status bijwerken
    taak_status = 'Voltooid' if actie == 'indienen' else 'In behandeling'
    conn.execute('UPDATE taken SET status=?, bijgewerkt=CURRENT_TIMESTAMP WHERE id=?',
                 (taak_status, taak_id))

    # Vrijwilliger-status bijwerken op basis van alle taken
    if actie == 'indienen':
        vw_id = taak['vrijwilliger_id']
        totaal   = conn.scalar('SELECT COUNT(*) FROM taken WHERE vrijwilliger_id = ?', (vw_id,))
        voltooid = conn.scalar(
            "SELECT COUNT(*) FROM taken WHERE vrijwilliger_id = ? AND status = 'Voltooid'",
            (vw_id,)
        )
        # Alle intakes voltooid → Actief, anders In behandeling
        nieuwe_status = 'Actief' if (totaal > 0 and voltooid >= totaal) else 'In behandeling'
        conn.execute(
            'UPDATE vrijwilligers SET status_vrijwilliger = ? WHERE id = ?',
            (nieuwe_status, vw_id)
        )

    conn.commit()
    conn.close()
    flash('Intake opgeslagen.' if actie == 'opslaan' else 'Intake ingediend.', 'success')
    return redirect(url_for('taken'))


# ── Beheer ─────────────────────────────────────────────────────────────────────

@app.route('/beheer')
@login_required
@rol_vereist('beheerder')
def beheer():
    conn = get_db()
    n_profielen = conn.scalar('SELECT COUNT(*) FROM profielen')
    n_eigenaren = conn.scalar('SELECT COUNT(*) FROM eigenaren')
    n_gebruikers = conn.scalar('SELECT COUNT(*) FROM gebruikers')
    conn.close()
    return render_template('beheer.html', n_profielen=n_profielen,
                           n_eigenaren=n_eigenaren, n_gebruikers=n_gebruikers)


# ── Eigenaren (beheer) ─────────────────────────────────────────────────────────

@app.route('/beheer/eigenaren')
@login_required
@rol_vereist('beheerder')
def eigenaren():
    zoek = request.args.get('q','').strip()
    prefill_id = request.args.get('prefill','').strip()
    conn = get_db()
    eigenaren_list = conn.execute(
        "SELECT e.*, g.voornaam||' '||g.achternaam AS gebruiker_naam FROM eigenaren e LEFT JOIN gebruikers g ON e.gebruiker_id = g.id ORDER BY e.achternaam, e.voornaam"
    ).fetchall()
    zoekresultaten = []
    if zoek:
        q = f'%{zoek}%'
        zoekresultaten = conn.execute(
            'SELECT * FROM vrijwilligers WHERE naam LIKE ? OR voornaam LIKE ? OR achternaam LIKE ? ORDER BY achternaam, voornaam, naam LIMIT 10',
            (q,q,q)
        ).fetchall()
    prefill = None
    if prefill_id:
        prefill = conn.execute('SELECT * FROM vrijwilligers WHERE id = ?', (prefill_id,)).fetchone()
    conn.close()
    return render_template('eigenaren.html', eigenaren=eigenaren_list,
                           zoekresultaten=zoekresultaten, zoek=zoek, prefill=prefill)


@app.route('/beheer/eigenaren/toevoegen', methods=['POST'])
@login_required
@rol_vereist('beheerder')
def eigenaar_toevoegen():
    voornaam = request.form.get('voornaam','').strip()
    achternaam = request.form.get('achternaam','').strip()
    if not voornaam or not achternaam:
        return redirect(url_for('eigenaren'))
    conn = get_db()
    conn.execute(
        'INSERT INTO eigenaren (voornaam, tussenvoegsel, achternaam, email, telefoonnummer) VALUES (?,?,?,?,?)',
        (voornaam, request.form.get('tussenvoegsel','').strip(), achternaam,
         request.form.get('email','').strip(), request.form.get('telefoonnummer','').strip())
    )
    conn.commit()
    conn.close()
    return redirect(url_for('eigenaren'))


@app.route('/beheer/eigenaren/verwijderen/<int:eid>', methods=['POST'])
@login_required
@rol_vereist('beheerder')
def eigenaar_verwijderen(eid):
    conn = get_db()
    conn.execute('DELETE FROM eigenaren WHERE id = ?', (eid,))
    conn.commit()
    conn.close()
    return redirect(url_for('eigenaren'))


@app.route('/beheer/eigenaren/<int:eid>/account-aanmaken', methods=['POST'])
@login_required
@rol_vereist('beheerder')
def eigenaar_account_aanmaken(eid):
    conn = get_db()
    eigenaar = conn.execute('SELECT * FROM eigenaren WHERE id = ?', (eid,)).fetchone()
    if not eigenaar or not eigenaar['email']:
        flash('Eigenaar heeft geen e-mailadres — account aanmaken niet mogelijk.', 'error')
        conn.close()
        return redirect(url_for('eigenaren'))

    email = eigenaar['email'].strip().lower()
    naam = f"{eigenaar['voornaam']} {eigenaar['achternaam']}"

    bestaand = conn.execute(
        'SELECT id FROM gebruikers WHERE LOWER(email) = ?', (email,)
    ).fetchone()

    if bestaand:
        # Koppel bestaand account aan eigenaar
        conn.execute('UPDATE eigenaren SET gebruiker_id = ? WHERE id = ?',
                     (bestaand['id'], eid))
        conn.commit()
        conn.close()
        flash(f'Bestaand account ({email}) gekoppeld aan {naam}.', 'info')
        return redirect(url_for('eigenaren'))

    # Nieuw account aanmaken
    ww = secrets.token_urlsafe(10)
    uid = conn.insert(
        'INSERT INTO gebruikers (voornaam, achternaam, email, wachtwoord) VALUES (?,?,?,?)',
        (eigenaar['voornaam'], eigenaar['achternaam'], email, generate_password_hash(ww))
    )
    conn.execute('INSERT INTO gebruiker_rollen (gebruiker_id, rol) VALUES (?,?)',
                 (uid, 'roleigenaar'))
    conn.execute('UPDATE eigenaren SET gebruiker_id = ? WHERE id = ?', (uid, eid))
    conn.commit()
    conn.close()

    email_welkom(email, naam, ww)
    flash(f'Account aangemaakt voor {naam} — welkomstmail verstuurd naar {email}.', 'success')
    return redirect(url_for('eigenaren'))


# ── Profielen (beheer) ─────────────────────────────────────────────────────────

@app.route('/beheer/profielen')
@login_required
@rol_vereist('beheerder')
def profielen_beheer():
    conn = get_db()
    profielen = conn.execute('''
        SELECT p.id, p.naam, p.eigenaar_id,
               p.vog_nodig, p.gedragscode_vereist, p.avg_akkoord_vereist,
               p.tweede_eigenaar_id, p.tweede_eigenaar_actief,
               e.voornaam||COALESCE(' '||NULLIF(e.tussenvoegsel,''),'')||' '||COALESCE(e.achternaam,'') AS eigenaar_naam,
               e2.voornaam||COALESCE(' '||NULLIF(e2.tussenvoegsel,''),'')||' '||COALESCE(e2.achternaam,'') AS tweede_eigenaar_naam,
               e2.email AS tweede_eigenaar_email
        FROM profielen p
        LEFT JOIN eigenaren e  ON p.eigenaar_id         = e.id
        LEFT JOIN eigenaren e2 ON p.tweede_eigenaar_id  = e2.id
        ORDER BY p.naam
    ''').fetchall()
    eigenaren = conn.execute('SELECT * FROM eigenaren ORDER BY achternaam, voornaam').fetchall()
    conn.close()
    return render_template('profielen_beheer.html', profielen=profielen, eigenaren=eigenaren)


@app.route('/beheer/profielen/toevoegen', methods=['POST'])
@login_required
@rol_vereist('beheerder')
def profiel_toevoegen():
    naam = request.form.get('naam','').strip()
    if not naam:
        return redirect(url_for('profielen_beheer'))
    conn = get_db()
    try:
        conn.execute('INSERT INTO profielen (naam, eigenaar_id) VALUES (?,?)',
                     (naam, request.form.get('eigenaar_id') or None))
        conn.commit()
    except Exception:
        conn.rollback()
        flash('Een profiel met deze naam bestaat al.', 'error')
    conn.close()
    return redirect(url_for('profielen_beheer'))


@app.route('/beheer/profielen/eigenaar/<int:pid>', methods=['POST'])
@login_required
@rol_vereist('beheerder')
def profiel_eigenaar(pid):
    conn = get_db()
    conn.execute('UPDATE profielen SET eigenaar_id=? WHERE id=?',
                 (request.form.get('eigenaar_id') or None, pid))
    conn.commit()
    conn.close()
    return redirect(url_for('profielen_beheer'))


@app.route('/beheer/profielen/vog/<int:pid>', methods=['POST'])
@login_required
@rol_vereist('beheerder')
def profiel_vog(pid):
    conn = get_db()
    conn.execute(
        'UPDATE profielen SET vog_nodig=?, gedragscode_vereist=?, avg_akkoord_vereist=? WHERE id=?',
        (request.form.get('vog_nodig',''),
         request.form.get('gedragscode_vereist',''),
         request.form.get('avg_akkoord_vereist',''),
         pid)
    )
    conn.commit()
    conn.close()
    flash('VOG-instellingen opgeslagen.', 'success')
    return redirect(url_for('profielen_beheer'))


@app.route('/beheer/profielen/tweede-eigenaar/<int:pid>', methods=['POST'])
@login_required
@rol_vereist('beheerder')
def profiel_tweede_eigenaar(pid):
    tweede_id = request.form.get('tweede_eigenaar_id') or None
    actief = 1 if request.form.get('tweede_eigenaar_actief') else 0
    conn = get_db()
    conn.execute(
        'UPDATE profielen SET tweede_eigenaar_id=?, tweede_eigenaar_actief=? WHERE id=?',
        (tweede_id, actief, pid)
    )
    conn.commit()
    conn.close()
    flash('Tweede bevoegde opgeslagen.', 'success')
    return redirect(url_for('profielen_beheer'))


@app.route('/beheer/profielen/verwijderen/<int:pid>', methods=['POST'])
@login_required
@rol_vereist('beheerder')
def profiel_verwijderen(pid):
    conn = get_db()
    conn.execute('DELETE FROM profielen WHERE id=?', (pid,))
    conn.commit()
    conn.close()
    return redirect(url_for('profielen_beheer'))


# ── Gebruikersbeheer ───────────────────────────────────────────────────────────

@app.route('/beheer/gebruikers')
@login_required
@rol_vereist('beheerder')
def gebruikers():
    conn = get_db()
    users = conn.execute('SELECT * FROM gebruikers ORDER BY achternaam, voornaam').fetchall()
    user_rollen = {}
    for u in users:
        user_rollen[u['id']] = [r['rol'] for r in conn.execute(
            'SELECT rol FROM gebruiker_rollen WHERE gebruiker_id = ?', (u['id'],)
        ).fetchall()]
    eigenaren_list = conn.execute(
        'SELECT * FROM eigenaren ORDER BY achternaam, voornaam'
    ).fetchall()
    conn.close()
    return render_template('gebruikers.html', gebruikers=users,
                           user_rollen=user_rollen, eigenaren=eigenaren_list,
                           alle_rollen=ALLE_ROLLEN)


@app.route('/beheer/gebruikers/toevoegen', methods=['POST'])
@login_required
@rol_vereist('beheerder')
def gebruiker_toevoegen():
    voornaam = request.form.get('voornaam','').strip()
    achternaam = request.form.get('achternaam','').strip()
    email = request.form.get('email','').strip().lower()
    wachtwoord = request.form.get('wachtwoord','')
    if not all([voornaam, achternaam, email, wachtwoord]):
        flash('Vul alle verplichte velden in.', 'error')
        return redirect(url_for('gebruikers'))
    conn = get_db()
    try:
        uid = conn.insert(
            'INSERT INTO gebruikers (voornaam, achternaam, email, wachtwoord) VALUES (?,?,?,?)',
            (voornaam, achternaam, email, generate_password_hash(wachtwoord))
        )
        for rol in request.form.getlist('rollen'):
            if rol in ALLE_ROLLEN:
                conn.execute('INSERT OR IGNORE INTO gebruiker_rollen (gebruiker_id, rol) VALUES (?,?)', (uid, rol))
        conn.commit()
        flash(f'Gebruiker {voornaam} {achternaam} aangemaakt.', 'success')
    except Exception:
        conn.rollback()
        flash('Dit e-mailadres is al in gebruik.', 'error')
    conn.close()
    return redirect(url_for('gebruikers'))


@app.route('/beheer/gebruikers/<int:uid>/rollen', methods=['POST'])
@login_required
@rol_vereist('beheerder')
def gebruiker_rollen_update(uid):
    nieuwe_rollen = [r for r in request.form.getlist('rollen') if r in ALLE_ROLLEN]
    conn = get_db()
    conn.execute('DELETE FROM gebruiker_rollen WHERE gebruiker_id=?', (uid,))
    for rol in nieuwe_rollen:
        conn.execute('INSERT INTO gebruiker_rollen (gebruiker_id, rol) VALUES (?,?)', (uid, rol))
    # Eigenaar koppelen
    eigenaar_id = request.form.get('eigenaar_id') or None
    conn.execute('UPDATE eigenaren SET gebruiker_id=NULL WHERE gebruiker_id=?', (uid,))
    if eigenaar_id:
        conn.execute('UPDATE eigenaren SET gebruiker_id=? WHERE id=?', (uid, eigenaar_id))
    conn.commit()
    conn.close()
    flash('Rollen bijgewerkt.', 'success')
    return redirect(url_for('gebruikers'))


@app.route('/beheer/gebruikers/<int:uid>/wachtwoord', methods=['POST'])
@login_required
@rol_vereist('beheerder')
def gebruiker_wachtwoord(uid):
    nieuw = request.form.get('wachtwoord','')
    if len(nieuw) < 6:
        flash('Wachtwoord moet minimaal 6 tekens zijn.', 'error')
        return redirect(url_for('gebruikers'))
    conn = get_db()
    conn.execute('UPDATE gebruikers SET wachtwoord=? WHERE id=?',
                 (generate_password_hash(nieuw), uid))
    conn.commit()
    conn.close()
    flash('Wachtwoord gewijzigd.', 'success')
    return redirect(url_for('gebruikers'))


@app.route('/beheer/gebruikers/verwijderen/<int:uid>', methods=['POST'])
@login_required
@rol_vereist('beheerder')
def gebruiker_verwijderen(uid):
    if uid == session.get('user_id'):
        flash('Je kunt je eigen account niet verwijderen.', 'error')
        return redirect(url_for('gebruikers'))
    conn = get_db()
    conn.execute('DELETE FROM gebruikers WHERE id=?', (uid,))
    conn.commit()
    conn.close()
    flash('Gebruiker verwijderd.', 'success')
    return redirect(url_for('gebruikers'))


# Initialiseer database bij opstarten (ook op Vercel)
try:
    init_db()
except Exception as _e:
    print(f'[init_db] {_e}')

if __name__ == '__main__':
    print('  App draait op: http://localhost:5000')
    debug = os.environ.get('FLASK_ENV') != 'production'
    app.run(debug=debug, port=int(os.environ.get('PORT', 5000)))
